from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import cv2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image


DEFAULT_ROOT = Path(__file__).resolve().parent
SHOT_DIR = "01_shots"
ASSET_DIR = "02_assets"

VIDEO_EXTENSIONS = {".mov", ".mp4", ".avi", ".mkv", ".mxf"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}
THREE_D_EXTENSIONS = {".abc", ".fbx", ".obj", ".ma", ".mb", ".blend", ".usd", ".usda", ".usdc"}
MOCAP_EXTENSIONS = {".bvh", ".c3d"}
PROCESS_ID_PATTERN = re.compile(r"^F\d{3}_", re.I)
ASSET_ID_PATTERN = re.compile(r"^A\d{3}_", re.I)

STAGE_NAMES = {
    "io": "原始素材",
    "原始素材": "原始素材",
    "mod": "Model（模型）",
    "previs": "Previsualization（预演）",
    "dmt": "Matte Painting（数字绘景）",
    "mg": "Motion Graphics（动态图形）",
    "mm": "Matchmove（跟踪）",
    "ani": "Animation（动画）",
    "env": "Environment（环境）",
    "lgt": "Lighting & Rendering（灯光渲染）",
    "efx": "Effects FX（特效）",
    "cfx": "Character FX（角色特效）",
    "roto": "Rotoscoping（抠像）",
    "paint": "Paint（擦除/修复）",
    "layout": "Layout（布局）",
    "cmp": "Compositing（合成）",
    "lfx": "最终文件（添加光效）",
}

STAGE_ORDER = {
    "io": 0,
    "原始素材": 0,
    "mod": 10,
    "previs": 20,
    "dmt": 30,
    "mg": 40,
    "mm": 50,
    "ani": 60,
    "env": 70,
    "lgt": 80,
    "efx": 90,
    "cfx": 100,
    "roto": 110,
    "paint": 120,
    "layout": 130,
    "cmp": 900,
    "lfx": 950,
}

ASSET_STYLE_NAMES = {
    "efx": {
        "fog": "雾效",
        "rain": "雨效",
        "fire": "火效",
        "dust": "尘效",
        "cloud": "云效",
        "smoke": "烟效",
        "particle": "粒子",
        "blood": "血效",
        "flare": "光晕",
    },
    "lgt": {
        "character": "角色",
        "prop": "道具",
        "set": "陈设",
        "env": "环境"
    },
    "dmt": {
        "env": "自然环境",
        "building": "人文建筑",
        "street": "街区",
        "sky": "天空",
        "decal": "纹理",
    },
}


@dataclass(frozen=True)
class ShotFile:
    file_id: str
    path: Path
    stage_code: str
    version: str
    is_final: bool


@dataclass(frozen=True)
class AssetFile:
    asset_id: str
    path: Path


@dataclass(frozen=True)
class RenamePlan:
    source: Path
    target: Path


def natural_key(path: Path) -> tuple:
    return tuple(int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", path.name))


def parse_stage_and_version(path: Path, shot_id: str) -> tuple[str, str]:
    stem = path.stem
    stem = re.sub(r"^[FA]\d{3}_", "", stem, flags=re.I)
    remainder = stem
    if stem.lower().startswith(shot_id.lower()):
        remainder = stem[len(shot_id) :].lstrip("_")
    parts = remainder.split("_")

    version = next((part for part in reversed(parts) if re.fullmatch(r"v\d+", part, re.I)), "")
    stage_code = ""
    for part in parts:
        normalized = part.lower()
        if normalized in STAGE_NAMES:
            stage_code = normalized
            break

    if not stage_code:
        stage_parts = [part for part in parts if not re.fullmatch(r"v\d+", part, re.I)]
        stage_code = stage_parts[0] if stage_parts else ""

    return stage_code.lower(), version.lower()


def stage_descriptors(path: Path, shot_id: str, stage_code: str) -> list[str]:
    stem = path.stem
    stem = re.sub(r"^[FA]\d{3}_", "", stem, flags=re.I)
    remainder = stem
    if stem.lower().startswith(shot_id.lower()):
        remainder = stem[len(shot_id) :].lstrip("_")

    parts = [part.lower() for part in remainder.split("_")]
    stage_code = stage_code.lower()
    for index, part in enumerate(parts):
        if part != stage_code:
            continue
        descriptors = []
        for descriptor in parts[index + 1 :]:
            if re.fullmatch(r"v\d+", descriptor, re.I):
                break
            descriptors.append(descriptor)
        return descriptors

    return []


def is_reposition_cmp(path: Path, shot_id: str, stage_code: str) -> bool:
    return stage_code == "cmp" and "reposition" in stage_descriptors(path, shot_id, stage_code)


def parse_asset_stage_and_style(path: Path) -> tuple[str, str]:
    stem = strip_id_prefix(path.stem)
    parts = [part.lower() for part in stem.split("_")]

    for index, part in enumerate(parts):
        if part not in STAGE_NAMES:
            continue

        style_names = ASSET_STYLE_NAMES.get(part, {})
        style = ""
        for descriptor in parts[index + 1 :]:
            if re.fullmatch(r"v\d+", descriptor, re.I):
                break
            if descriptor in style_names:
                style = style_names[descriptor]
                break
        return part, style

    return "", ""


def version_number(version: str) -> int:
    match = re.search(r"\d+", version)
    return int(match.group()) if match else 0


def stage_sort_key(item: tuple[Path, str, str]) -> tuple:
    path, stage_code, version = item
    return (STAGE_ORDER.get(stage_code, 500), stage_code, version_number(version), natural_key(path))


def strip_id_prefix(name: str) -> str:
    return re.sub(r"^[FA]\d{3}_", "", name, flags=re.I)


def normalized_output_suffix(path: Path) -> str:
    return ".mp4" if path.suffix.lower() == ".mov" else path.suffix


def prefixed_output_name(prefix: str, path: Path) -> str:
    stripped = strip_id_prefix(path.name)
    stripped_path = Path(stripped)
    return f"{prefix}_{stripped_path.stem}{normalized_output_suffix(path)}"


def is_excel_output(path: Path, shot_id: str) -> bool:
    return path.name == f"{shot_id}_数据说明.xlsx"


def list_process_paths(group_dir: Path) -> list[Path]:
    return sorted(
        [path for path in (group_dir / SHOT_DIR).glob("*") if path.is_file()],
        key=lambda path: stage_sort_key((path, *parse_stage_and_version(path, group_dir.name))),
    )


def list_asset_paths(group_dir: Path) -> list[Path]:
    return sorted(
        [path for path in (group_dir / ASSET_DIR).glob("*") if path.is_file()],
        key=natural_key,
    )


def build_rename_plans(group_dir: Path) -> list[RenamePlan]:
    plans: list[RenamePlan] = []

    for index, path in enumerate(list_process_paths(group_dir), start=1):
        target_name = prefixed_output_name(f"F{index:03d}", path)
        target = path.with_name(target_name)
        if path.name != target_name:
            plans.append(RenamePlan(path, target))

    for index, path in enumerate(list_asset_paths(group_dir), start=1):
        target_name = prefixed_output_name(f"A{index:03d}", path)
        target = path.with_name(target_name)
        if path.name != target_name:
            plans.append(RenamePlan(path, target))

    return plans


def validate_rename_plans(plans: list[RenamePlan]) -> None:
    targets = [plan.target.resolve() for plan in plans]
    duplicate_targets = {target for target in targets if targets.count(target) > 1}
    if duplicate_targets:
        duplicates = "\n".join(str(path) for path in sorted(duplicate_targets))
        raise ValueError(f"重命名目标存在重复，已停止执行：\n{duplicates}")

    sources = {plan.source.resolve() for plan in plans}
    conflicts = [
        plan.target
        for plan in plans
        if plan.target.exists() and plan.target.resolve() not in sources
    ]
    if conflicts:
        conflict_text = "\n".join(str(path) for path in conflicts)
        raise FileExistsError(f"重命名目标已存在，已停止执行：\n{conflict_text}")


def apply_rename_plans(plans: list[RenamePlan]) -> None:
    validate_rename_plans(plans)

    temp_plans = []
    for index, plan in enumerate(plans, start=1):
        temp = plan.source.with_name(f".rename_tmp_{index:03d}_{plan.source.name}")
        plan.source.rename(temp)
        temp_plans.append(RenamePlan(temp, plan.target))

    for plan in temp_plans:
        plan.source.rename(plan.target)


def normalize_group_filenames(group_dir: Path) -> list[RenamePlan]:
    plans = build_rename_plans(group_dir)
    if plans:
        apply_rename_plans(plans)
    return plans


def file_size_mb(path: Path) -> float:
    return round(path.stat().st_size / 1024 / 1024, 3)


def relative_path(path: Path, base: Path) -> str:
    return path.relative_to(base).as_posix()


def video_metadata(path: Path) -> tuple[str, float]:
    capture = cv2.VideoCapture(str(path))
    if not capture.isOpened():
        return "", 0.0

    width = int(capture.get(cv2.CAP_PROP_FRAME_WIDTH) or 0)
    height = int(capture.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0)
    frames = capture.get(cv2.CAP_PROP_FRAME_COUNT) or 0
    fps = capture.get(cv2.CAP_PROP_FPS) or 0
    capture.release()

    resolution = f"{width}*{height}" if width and height else ""
    duration = round(frames / fps, 3) if fps else 0.0
    return resolution, duration


def image_resolution(path: Path) -> str:
    with Image.open(path) as image:
        return f"{image.width}*{image.height}"


def asset_format(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in IMAGE_EXTENSIONS:
        return "图片"
    if suffix in VIDEO_EXTENSIONS:
        return "渲染视频"
    if suffix in THREE_D_EXTENSIONS:
        return "三维源文件"
    if suffix in MOCAP_EXTENSIONS:
        return "动捕数据"
    return "其它格式数据"


def collect_shot_files(group_dir: Path, shot_id: str) -> list[ShotFile]:
    shot_paths = list_process_paths(group_dir)
    parsed = [(path, *parse_stage_and_version(path, shot_id)) for path in shot_paths]
    parsed = sorted(parsed, key=stage_sort_key)
    cmp_versions = [
        version_number(version)
        for path, stage, version in parsed
        if stage == "cmp" and not is_reposition_cmp(path, shot_id, stage)
    ]
    final_cmp_version = max(cmp_versions) if cmp_versions else None

    shot_files = []
    for index, (path, stage_code, version) in enumerate(parsed, start=1):
        id_match = re.match(r"^(F\d{3})_", path.name, flags=re.I)
        is_final = (
            stage_code == "cmp"
            and not is_reposition_cmp(path, shot_id, stage_code)
            and version_number(version) == final_cmp_version
        )
        shot_files.append(
            ShotFile(
                file_id=id_match.group(1).upper() if id_match else f"F{index:03d}",
                path=path,
                stage_code=stage_code,
                version=version,
                is_final=is_final,
            )
        )
    return shot_files


def collect_asset_files(group_dir: Path) -> list[AssetFile]:
    asset_paths = list_asset_paths(group_dir)
    asset_files = []
    for index, path in enumerate(asset_paths, start=1):
        id_match = re.match(r"^(A\d{3})_", path.name, flags=re.I)
        asset_files.append(AssetFile(asset_id=id_match.group(1).upper() if id_match else f"A{index:03d}", path=path))
    return asset_files


def build_sheet1_rows(group_dir: Path, shot_files: Iterable[ShotFile], root: Path) -> list[dict]:
    rows = []
    for shot_file in shot_files:
        resolution, duration = video_metadata(shot_file.path)
        rows.append(
            {
                "file_id": shot_file.file_id,
                "file_name": shot_file.path.name,
                "file_resolution": resolution,
                "file_duration": duration,
                "file_size": file_size_mb(shot_file.path),
                "is_final": "是" if shot_file.is_final else "否",
                "is_log": "否" if shot_file.is_final else "",
                "graded_file_id": "",
                "file_path": relative_path(shot_file.path, root),
            }
        )
    return rows


def build_sheet2_rows(group_dir: Path, asset_files: Iterable[AssetFile], root: Path) -> list[dict]:
    rows = []
    for asset_file in asset_files:
        stage_code, style = parse_asset_stage_and_style(asset_file.path)
        rows.append(
            {
                "asset_id": asset_file.asset_id,
                "asset_name": asset_file.path.name,
                "asset_type": STAGE_NAMES.get(stage_code, ""),
                "asset_format": asset_format(asset_file.path),
                "view_angle": "成片机位",
                "style": style,
                "file_path": relative_path(asset_file.path, root),
            }
        )
    return rows


def build_sheet3_rows(shot_files: list[ShotFile], asset_files: list[AssetFile]) -> list[dict]:
    rows = []
    previous_output_id = ""
    asset_ids = ",".join(asset_file.asset_id for asset_file in asset_files)
    stage_context: dict[str, tuple[str, str]] = {}
    next_step_number = 1

    for index, shot_file in enumerate(shot_files, start=1):
        stage_name = STAGE_NAMES.get(shot_file.stage_code, shot_file.stage_code)
        if index == 1:
            step_id = f"{next_step_number:03d}"
            input_file_id = ""
            ref_asset_ids = ""
            modification = "作为本组镜头的原始输入素材。"
            stage_context[shot_file.stage_code] = (step_id, input_file_id)
            next_step_number += 1
        else:
            if shot_file.stage_code in stage_context:
                step_id, input_file_id = stage_context[shot_file.stage_code]
            else:
                step_id = f"{next_step_number:03d}"
                input_file_id = previous_output_id
                stage_context[shot_file.stage_code] = (step_id, input_file_id)
                next_step_number += 1

            ref_asset_ids = asset_ids if shot_file.stage_code == "cmp" else ""
            modification = f"基于 {input_file_id} 进行{stage_name}处理，输出 {shot_file.file_id}。"
            if ref_asset_ids:
                modification = f"引用素材：{ref_asset_ids}。" + modification
            if shot_file.is_final:
                modification = f"基于 {input_file_id} 完成最终合成输出，作为本组最终成片。"
                if ref_asset_ids:
                    modification = f"引用素材：{ref_asset_ids}。" + modification

        rows.append(
            {
                "step_id": step_id,
                "stage_code": shot_file.stage_code,
                "stage_name": stage_name,
                "stage_version": shot_file.version or "无",
                "input_file_id": input_file_id,
                "ref_asset_ids": ref_asset_ids,
                "modification_plain": modification,
                "output_file_id": shot_file.file_id,
            }
        )
        previous_output_id = shot_file.file_id
    return rows


def write_rows(ws, headers: list[str], rows: list[dict]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def generate_group_excel(group_dir: Path, root: Path) -> Path:
    shot_id = group_dir.name
    shot_files = collect_shot_files(group_dir, shot_id)
    asset_files = collect_asset_files(group_dir)

    if not shot_files:
        raise ValueError(f"{group_dir} 缺少 {SHOT_DIR} 文件，无法生成数据说明。")

    workbook = Workbook()
    workbook.remove(workbook.active)

    sheet1 = workbook.create_sheet("Sheet1_过程文件清单")
    write_rows(
        sheet1,
        [
            "file_id",
            "file_name",
            "file_resolution",
            "file_duration",
            "file_size",
            "is_final",
            "is_log",
            "graded_file_id",
            "file_path",
        ],
        build_sheet1_rows(group_dir, shot_files, root),
    )

    sheet2 = workbook.create_sheet("Sheet2_素材文件清单")
    write_rows(
        sheet2,
        [
            "asset_id",
            "asset_name",
            "asset_type",
            "asset_format",
            "view_angle",
            "style",
            "file_path",
        ],
        build_sheet2_rows(group_dir, asset_files, root),
    )

    sheet3 = workbook.create_sheet("Sheet3_编辑关系与修改说明")
    write_rows(
        sheet3,
        [
            "step_id",
            "stage_code",
            "stage_name",
            "stage_version",
            "input_file_id",
            "ref_asset_ids",
            "modification_plain",
            "output_file_id",
        ],
        build_sheet3_rows(shot_files, asset_files),
    )

    output_path = group_dir / f"{shot_id}_数据说明.xlsx"
    try:
        workbook.save(output_path)
    except PermissionError as exc:
        raise PermissionError(f"无法写入 {output_path}，请先关闭已打开的 Excel/WPS 文件后重试。") from exc
    return output_path


def iter_group_dirs(root: Path) -> Iterable[Path]:
    for path in sorted(root.iterdir(), key=natural_key):
        if not path.is_dir() or path.name.startswith("."):
            continue
        if (path / SHOT_DIR).is_dir() and (path / ASSET_DIR).is_dir():
            yield path


def resolve_root(argv: list[str]) -> Path:
    if len(argv) > 1:
        return Path(argv[1]).expanduser().resolve()
    return DEFAULT_ROOT


def main(argv: list[str] | None = None) -> None:
    argv = argv or sys.argv
    root = resolve_root(argv)
    if not root.is_dir():
        raise SystemExit(f"数据根目录不存在: {root}")

    output_paths = []
    skipped_empty_groups = []
    for group_dir in iter_group_dirs(root):
        if not list_process_paths(group_dir):
            print(f"跳过空数据组: {group_dir}")
            skipped_empty_groups.append(group_dir)
            continue
        rename_plans = normalize_group_filenames(group_dir)
        for plan in rename_plans:
            print(f"已重命名: {plan.source.name} -> {plan.target.name}")
        output_paths.append(generate_group_excel(group_dir, root))

    if not output_paths:
        if skipped_empty_groups:
            print("没有非空数据组，未生成 Excel。")
            return
        raise SystemExit("未找到包含 01_shots 和 02_assets 的数据组目录。")

    for output_path in output_paths:
        print(f"已生成: {output_path}")


if __name__ == "__main__":
    main()
